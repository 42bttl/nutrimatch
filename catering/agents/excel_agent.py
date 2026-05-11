from __future__ import annotations

import os
from datetime import date
from typing import List

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from catering.data.allergen_codes import ALLERGEN_CODES, CIRCLED_NUMBERS, KOREAN_16_ALLERGENS
from catering.schemas import CateringPlanResult

# ─── Style constants ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUB_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ACCENT_FILL = PatternFill("solid", fgColor="DEEAF1")
RED_FILL = PatternFill("solid", fgColor="FFCCCC")
GREEN_FILL = PatternFill("solid", fgColor="CCFFCC")
YELLOW_FILL = PatternFill("solid", fgColor="FFFFCC")
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
CCP_BORDER = Border(
    left=Side(style="medium", color="FF0000"),
    right=Side(style="medium", color="FF0000"),
    top=Side(style="medium", color="FF0000"),
    bottom=Side(style="medium", color="FF0000"),
)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _hdr(ws, row, col, value, fill=None, font=None, border=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if border:
        cell.border = border or THIN_BORDER
    if align:
        cell.alignment = align
    return cell


def _style_row(ws, row, cols, fill=None, border=True):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        if fill:
            cell.fill = fill
        if border:
            cell.border = THIN_BORDER
        cell.alignment = CENTER


# ─── Main Excel Agent ─────────────────────────────────────────────────────────

class ExcelAgent:

    def render(self, plan: CateringPlanResult, output_dir: str) -> str:
        os.makedirs(output_dir, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self._sheet_monthly_overview(wb, plan)
        self._sheet_breakfast(wb, plan)
        self._sheet_lunch(wb, plan)
        self._sheet_dinner(wb, plan)
        self._sheet_night_snack(wb, plan)
        self._sheet_nutrition(wb, plan)
        self._sheet_allergens(wb, plan)
        self._sheet_cost(wb, plan)
        self._sheet_procurement(wb, plan)
        self._sheet_recipes(wb, plan)

        filename = f"{plan.ship}_{plan.year}{plan.month:02d}.xlsx"
        path = os.path.join(output_dir, filename)
        wb.save(path)
        return path

    # ── Sheet 1: 월간 식단표 ────────────────────────────────────────────────

    def _sheet_monthly_overview(self, wb, plan: CateringPlanResult):
        ws = wb.create_sheet("월간 식단표")
        ws.sheet_view.showGridLines = False

        title = f"{plan.year}년 {plan.month}월 월간 식단표 — {plan.ship_name_ko} ({plan.headcount}명)"
        ws.merge_cells("A1:H1")
        cell = ws["A1"]
        cell.value = title
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=14)
        cell.alignment = CENTER
        ws.row_dimensions[1].height = 30

        headers = ["날짜", "요일", "조식 (주찬)", "중식 (주찬1)", "중식 (주찬2)", "석식 (주찬)", "국", "야식"]
        for col, h in enumerate(headers, 1):
            _hdr(ws, 2, col, h, SUB_HEADER_FILL, SUB_HEADER_FONT, THIN_BORDER, CENTER)

        menu_index = {}
        try:
            import pandas as pd
            from pathlib import Path
            data_dir = Path(__file__).parent.parent / "data"
            menu_df = pd.read_csv(data_dir / "menu_db.csv").set_index("menu_id")
            menu_index = menu_df["menu_name_ko"].to_dict()
        except Exception:
            pass

        WEEKDAYS_KO = ["월", "화", "수", "목", "금", "토", "일"]

        for row_idx, day in enumerate(plan.daily_menus, 3):
            ws.row_dimensions[row_idx].height = 30
            ws.cell(row=row_idx, column=1, value=day.date.strftime("%m/%d")).alignment = CENTER
            ws.cell(row=row_idx, column=1).border = THIN_BORDER
            wd = WEEKDAYS_KO[day.date.weekday()]
            wd_cell = ws.cell(row=row_idx, column=2, value=wd)
            wd_cell.alignment = CENTER
            wd_cell.border = THIN_BORDER
            if wd in ("토", "일"):
                wd_cell.fill = ACCENT_FILL

            bf_main = menu_index.get(day.breakfast.main_dish_ids[0] if day.breakfast.main_dish_ids else "", "")
            lu_main1 = menu_index.get(day.lunch.main_dish_ids[0] if day.lunch.main_dish_ids else "", "")
            lu_main2 = menu_index.get(day.lunch.main_dish_ids[1] if len(day.lunch.main_dish_ids) > 1 else "", "")
            di_main = menu_index.get(day.dinner.main_dish_ids[0] if day.dinner.main_dish_ids else "", "")
            lu_soup = menu_index.get(day.lunch.soup_id, "")

            for col, val in enumerate([bf_main, lu_main1, lu_main2, di_main, lu_soup, "삶은계란/바나나/우유"], 3):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = CENTER
                cell.border = THIN_BORDER

        for col_width in [(1, 8), (2, 5), (3, 16), (4, 16), (5, 16), (6, 16), (7, 16), (8, 18)]:
            ws.column_dimensions[get_column_letter(col_width[0])].width = col_width[1]

    # ── Sheet 2: 조식 ───────────────────────────────────────────────────────

    def _sheet_breakfast(self, wb, plan: CateringPlanResult):
        ws = wb.create_sheet("조식")
        menu_index = self._load_menu_index()
        headers = ["날짜", "밥", "국", "주찬", "주부찬", "부찬1", "부찬2", "후식"]
        self._write_meal_sheet(ws, plan, "조식", "breakfast", headers, menu_index)

    # ── Sheet 3: 중식 ───────────────────────────────────────────────────────

    def _sheet_lunch(self, wb, plan: CateringPlanResult):
        ws = wb.create_sheet("중식")
        menu_index = self._load_menu_index()
        headers = ["날짜", "밥", "국", "주찬1", "주찬2", "주부찬1", "주부찬2", "부찬1", "부찬2", "과일", "생채소", "후식"]
        self._write_meal_sheet(ws, plan, "중식", "lunch", headers, menu_index)

    # ── Sheet 4: 석식 ───────────────────────────────────────────────────────

    def _sheet_dinner(self, wb, plan: CateringPlanResult):
        ws = wb.create_sheet("석식")
        menu_index = self._load_menu_index()
        headers = ["날짜", "밥", "국", "주찬", "주부찬1", "주부찬2", "부찬1", "부찬2", "후식"]
        self._write_meal_sheet(ws, plan, "석식", "dinner", headers, menu_index)

    def _write_meal_sheet(self, ws, plan, title_ko, meal_attr, headers, menu_index):
        ws.sheet_view.showGridLines = False
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        cell = ws["A1"]
        cell.value = f"{plan.year}년 {plan.month}월 {title_ko} 식단표 — {plan.ship_name_ko}"
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=13)
        cell.alignment = CENTER
        ws.row_dimensions[1].height = 25

        for col, h in enumerate(headers, 1):
            _hdr(ws, 2, col, h, SUB_HEADER_FILL, SUB_HEADER_FONT, THIN_BORDER, CENTER)

        for row_idx, day in enumerate(plan.daily_menus, 3):
            ws.row_dimensions[row_idx].height = 25
            slot = getattr(day, meal_attr)
            ws.cell(row=row_idx, column=1, value=day.date.strftime("%m/%d")).alignment = CENTER
            ws.cell(row=row_idx, column=1).border = THIN_BORDER

            values = [
                menu_index.get(slot.rice_id, slot.rice_id),
                menu_index.get(slot.soup_id, slot.soup_id),
            ]
            for mid in slot.main_dish_ids:
                values.append(menu_index.get(mid, mid))
            for mid in slot.sub_main_ids:
                values.append(menu_index.get(mid, mid))
            for mid in slot.side_dish_ids:
                values.append(menu_index.get(mid, mid))
            # Extras / 후식
            extras_str = " / ".join(slot.extras) if slot.extras else ""
            values.append(extras_str)

            for col, val in enumerate(values[:len(headers) - 1], 2):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = CENTER
                cell.border = THIN_BORDER

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 14

    # ── Sheet 5: 야식 ───────────────────────────────────────────────────────

    def _sheet_night_snack(self, wb, plan: CateringPlanResult):
        ws = wb.create_sheet("야식")
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:D1")
        ws["A1"].value = f"{plan.year}년 {plan.month}월 야식 — {plan.ship_name_ko}"
        ws["A1"].fill = HEADER_FILL
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].alignment = CENTER

        for col, h in enumerate(["날짜", "삶은계란", "바나나", "우유/두유"], 1):
            _hdr(ws, 2, col, h, SUB_HEADER_FILL, SUB_HEADER_FONT, THIN_BORDER, CENTER)

        for row_idx, day in enumerate(plan.daily_menus, 3):
            ws.cell(row=row_idx, column=1, value=day.date.strftime("%m/%d")).border = THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = CENTER
            for col, val in enumerate(["1개", "1개", "200ml"], 2):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = CENTER
                cell.border = THIN_BORDER

        for i, w in enumerate([8, 12, 12, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 6: 영양분석 ───────────────────────────────────────────────────

    def _sheet_nutrition(self, wb, plan: CateringPlanResult):
        ws = wb.create_sheet("영양분석")
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:I1")
        ws["A1"].value = f"{plan.year}년 {plan.month}월 영양분석 — {plan.ship_name_ko}"
        ws["A1"].fill = HEADER_FILL
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 25

        headers = ["날짜", "조식(kcal)", "중식(kcal)", "석식(kcal)", "야식(kcal)", "합계(kcal)", "단백질(g)", "탄수화물(g)", "지방(g)", "식이섬유(g)", "목표충족"]
        for col, h in enumerate(headers, 1):
            _hdr(ws, 2, col, h, SUB_HEADER_FILL, SUB_HEADER_FONT, THIN_BORDER, CENTER)

        for row_idx, nd in enumerate(plan.nutrition_analysis, 3):
            ws.row_dimensions[row_idx].height = 20
            values = [
                nd.date.strftime("%m/%d"),
                nd.breakfast_kcal,
                nd.lunch_kcal,
                nd.dinner_kcal,
                nd.snack_kcal,
                nd.total_kcal,
                nd.protein_g,
                nd.carb_g,
                nd.fat_g,
                nd.fiber_g,
                "✓" if nd.within_target else "✗",
            ]
            row_fill = GREEN_FILL if nd.within_target else RED_FILL
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = CENTER
                cell.border = THIN_BORDER
                if col == 6:  # total kcal column
                    cell.fill = row_fill

        # Summary row
        summary_row = len(plan.nutrition_analysis) + 3
        ws.cell(row=summary_row, column=1, value="평균").fill = ACCENT_FILL
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        ws.cell(row=summary_row, column=1).alignment = CENTER

        if plan.nutrition_analysis:
            avg_kcal = sum(n.total_kcal for n in plan.nutrition_analysis) / len(plan.nutrition_analysis)
            avg_protein = sum(n.protein_g for n in plan.nutrition_analysis) / len(plan.nutrition_analysis)
            avg_fiber = sum(n.fiber_g for n in plan.nutrition_analysis) / len(plan.nutrition_analysis)
            for col, val in zip([6, 7, 10], [round(avg_kcal, 1), round(avg_protein, 1), round(avg_fiber, 1)]):
                cell = ws.cell(row=summary_row, column=col, value=val)
                cell.font = Font(bold=True)
                cell.alignment = CENTER
                cell.border = THIN_BORDER
                cell.fill = ACCENT_FILL

        # Target reference row
        target_row = summary_row + 1
        ws.cell(row=target_row, column=1, value="목표 범위")
        ws.cell(row=target_row, column=6, value="2700~3200")
        ws.cell(row=target_row, column=7, value="90~120")
        ws.cell(row=target_row, column=10, value="25 이상")
        for col in [1, 6, 7, 10]:
            cell = ws.cell(row=target_row, column=col)
            cell.font = Font(bold=True, color="1F4E79")
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.fill = YELLOW_FILL

        # Bar chart for daily kcal
        if len(plan.nutrition_analysis) >= 2:
            chart = BarChart()
            chart.type = "col"
            chart.title = "일별 총 열량 (kcal)"
            chart.y_axis.title = "kcal"
            chart.x_axis.title = "날짜"
            chart.style = 10
            chart.width = 20
            chart.height = 12

            data_ref = Reference(ws, min_col=6, min_row=2, max_row=len(plan.nutrition_analysis) + 2)
            chart.add_data(data_ref, titles_from_data=True)
            dates_ref = Reference(ws, min_col=1, min_row=3, max_row=len(plan.nutrition_analysis) + 2)
            chart.set_categories(dates_ref)
            ws.add_chart(chart, f"A{target_row + 2}")

        for i, w in enumerate([8, 12, 12, 12, 12, 13, 12, 14, 12, 13, 10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 7: 알레르기표시 ───────────────────────────────────────────────

    def _sheet_allergens(self, wb, plan: CateringPlanResult):
        ws = wb.create_sheet("알레르기표시")
        ws.sheet_view.showGridLines = False

        header_text = (
            f"{plan.year}년 {plan.month}월 알레르기 표시 — {plan.ship_name_ko}  |  "
            "알레르기 코드: " + " ".join(
                f"{CIRCLED_NUMBERS[i]}{name}" for i, name in enumerate(KOREAN_16_ALLERGENS)
            )
        )
        col_count = 2 + len(KOREAN_16_ALLERGENS)
        ws.merge_cells(f"A1:{get_column_letter(col_count)}1")
        ws["A1"].value = header_text
        ws["A1"].fill = HEADER_FILL
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=10)
        ws["A1"].alignment = LEFT

        # Column headers
        _hdr(ws, 2, 1, "메뉴명", SUB_HEADER_FILL, SUB_HEADER_FONT, THIN_BORDER, CENTER)
        _hdr(ws, 2, 2, "알레르기 목록", SUB_HEADER_FILL, SUB_HEADER_FONT, THIN_BORDER, CENTER)
        for i, name in enumerate(KOREAN_16_ALLERGENS):
            code = CIRCLED_NUMBERS[i]
            _hdr(ws, 2, 3 + i, f"{code}{name}", SUB_HEADER_FILL, SUB_HEADER_FONT, THIN_BORDER, CENTER)

        for row_idx, ar in enumerate(plan.allergen_table, 3):
            ws.cell(row=row_idx, column=1, value=ar.menu_name).border = THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = LEFT

            if ar.allergens_present:
                codes = " ".join(
                    CIRCLED_NUMBERS[ALLERGEN_CODES[a] - 1]
                    for a in ar.allergens_present
                    if a in ALLERGEN_CODES
                )
                ws.cell(row=row_idx, column=2, value=codes).border = THIN_BORDER
                ws.cell(row=row_idx, column=2).alignment = CENTER
            else:
                ws.cell(row=row_idx, column=2, value="없음").border = THIN_BORDER
                ws.cell(row=row_idx, column=2).alignment = CENTER

            for i, allergen in enumerate(KOREAN_16_ALLERGENS):
                cell = ws.cell(row=row_idx, column=3 + i)
                cell.border = THIN_BORDER
                cell.alignment = CENTER
                if allergen in ar.allergens_present:
                    cell.value = "●"
                    cell.fill = ORANGE_FILL
                    cell.font = Font(bold=True, color="C00000")

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 20
        for i in range(len(KOREAN_16_ALLERGENS)):
            ws.column_dimensions[get_column_letter(3 + i)].width = 8

    # ── Sheet 8: 원가분석 ───────────────────────────────────────────────────

    def _sheet_cost(self, wb, plan: CateringPlanResult):
        ws = wb.create_sheet("원가분석")
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:G1")
        ws["A1"].value = f"{plan.year}년 {plan.month}월 원가분석 — {plan.ship_name_ko} ({plan.headcount}명)"
        ws["A1"].fill = HEADER_FILL
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].alignment = CENTER

        headers = ["날짜", "조식 원가", "중식 원가", "석식 원가", "야식 원가", "일일 합계", "1인당 원가"]
        for col, h in enumerate(headers, 1):
            _hdr(ws, 2, col, h, SUB_HEADER_FILL, SUB_HEADER_FONT, THIN_BORDER, CENTER)

        try:
            import pandas as pd
            from pathlib import Path
            data_dir = Path(__file__).parent.parent / "data"
            menu_df = pd.read_csv(data_dir / "menu_db.csv").set_index("menu_id")
            cost_index = menu_df["cost_per_serving_krw"].to_dict()
        except Exception:
            cost_index = {}

        NIGHT_SNACK_COST = 2000  # fixed
        CONDIMENT_COST = 300
        budget_per_meal = 8000

        running_total = 0
        for row_idx, day in enumerate(plan.daily_menus, 3):
            bf_cost = self._slot_cost(day.breakfast, cost_index) + CONDIMENT_COST
            lu_cost = self._slot_cost(day.lunch, cost_index) + CONDIMENT_COST
            di_cost = self._slot_cost(day.dinner, cost_index) + CONDIMENT_COST
            sn_cost = NIGHT_SNACK_COST
            daily_total = bf_cost + lu_cost + di_cost + sn_cost
            running_total += daily_total
            per_person = daily_total

            values = [
                day.date.strftime("%m/%d"),
                f"₩{bf_cost:,}",
                f"₩{lu_cost:,}",
                f"₩{di_cost:,}",
                f"₩{sn_cost:,}",
                f"₩{daily_total:,}",
                f"₩{per_person:,}",
            ]
            row_fill = RED_FILL if per_person > budget_per_meal * 4 else None
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = CENTER
                cell.border = THIN_BORDER
                if row_fill and col in (6, 7):
                    cell.fill = row_fill

        # Grand total
        total_row = len(plan.daily_menus) + 3
        ws.cell(row=total_row, column=1, value="합계").fill = ACCENT_FILL
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).alignment = CENTER
        ws.cell(row=total_row, column=6, value=f"₩{running_total:,}").fill = ACCENT_FILL
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        ws.cell(row=total_row, column=6).alignment = CENTER

        for i, w in enumerate([8, 12, 12, 12, 12, 14, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _slot_cost(self, slot, cost_index: dict) -> int:
        total = 0
        for mid in (
            [slot.rice_id, slot.soup_id]
            + slot.main_dish_ids
            + slot.sub_main_ids
            + slot.side_dish_ids
        ):
            total += int(cost_index.get(mid, 0))
        return total

    # ── Sheet 9: 발주량 집계 ────────────────────────────────────────────────

    def _sheet_procurement(self, wb, plan: CateringPlanResult):
        ws = wb.create_sheet("발주량 집계")
        ws.sheet_view.showGridLines = False
        ws.merge_cells("A1:H1")
        ws["A1"].value = f"{plan.year}년 {plan.month}월 발주량 집계 — {plan.ship_name_ko} ({plan.headcount}명, 14일)"
        ws["A1"].fill = HEADER_FILL
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].alignment = CENTER

        headers = ["분류", "품목명", "단위", "1인분(g)", "총량(g)", "총량(kg)", "단가(원/kg)", "총금액(원)"]
        for col, h in enumerate(headers, 1):
            _hdr(ws, 2, col, h, SUB_HEADER_FILL, SUB_HEADER_FONT, THIN_BORDER, CENTER)

        current_cat = None
        grand_total = 0
        for row_idx, line in enumerate(plan.procurement, 3):
            cat_fill = ACCENT_FILL if line.category != current_cat else None
            current_cat = line.category

            values = [
                line.category,
                line.ingredient,
                line.unit,
                line.per_serving_g,
                f"{line.total_g:,.0f}",
                f"{line.total_kg:,.2f}",
                f"₩{line.unit_price_krw:,}",
                f"₩{line.total_cost_krw:,}",
            ]
            grand_total += line.total_cost_krw
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.alignment = CENTER
                cell.border = THIN_BORDER
                if line.haccp_critical:
                    cell.fill = PatternFill("solid", fgColor="FFF2CC")
                elif cat_fill:
                    cell.fill = cat_fill

        total_row = len(plan.procurement) + 3
        ws.cell(row=total_row, column=1, value="합계").fill = ACCENT_FILL
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).alignment = CENTER
        ws.cell(row=total_row, column=8, value=f"₩{grand_total:,}").fill = ACCENT_FILL
        ws.cell(row=total_row, column=8).font = Font(bold=True)
        ws.cell(row=total_row, column=8).alignment = CENTER

        # Legend
        legend_row = total_row + 2
        ws.cell(row=legend_row, column=1, value="※ 노란색 음영: HACCP 중요 관리 식재료")
        ws.cell(row=legend_row, column=1).fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=legend_row, column=1).font = Font(bold=True, size=9)

        for i, w in enumerate([12, 16, 8, 10, 12, 10, 14, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 10: 조리지침서 ────────────────────────────────────────────────

    def _sheet_recipes(self, wb, plan: CateringPlanResult):
        ws = wb.create_sheet("조리지침서")
        ws.sheet_view.showGridLines = False

        ws.merge_cells("A1:E1")
        ws["A1"].value = f"{plan.year}년 {plan.month}월 조리지침서(SOP) — {plan.ship_name_ko} ({plan.headcount}명)"
        ws["A1"].fill = HEADER_FILL
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].alignment = CENTER

        current_row = 2
        for sop in plan.recipes:
            # Recipe header
            ws.merge_cells(f"A{current_row}:E{current_row}")
            hdr_cell = ws.cell(row=current_row, column=1, value=f"■ {sop.menu_name}")
            hdr_cell.fill = SUB_HEADER_FILL
            hdr_cell.font = Font(bold=True, color="FFFFFF", size=11)
            hdr_cell.alignment = LEFT
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            # Ingredients
            ws.cell(row=current_row, column=1, value="[재료]").font = Font(bold=True)
            ws.cell(row=current_row, column=1).fill = ACCENT_FILL
            ws.cell(row=current_row, column=1).border = THIN_BORDER
            ing_text = " | ".join(sop.ingredients)
            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws.cell(row=current_row, column=2, value=ing_text).border = THIN_BORDER
            ws.cell(row=current_row, column=2).alignment = LEFT
            current_row += 1

            # Cooking steps
            ws.cell(row=current_row, column=1, value="[조리 단계]").font = Font(bold=True)
            ws.cell(row=current_row, column=1).fill = ACCENT_FILL
            ws.cell(row=current_row, column=1).border = THIN_BORDER
            current_row += 1
            for step in sop.steps:
                ws.merge_cells(f"A{current_row}:E{current_row}")
                cell = ws.cell(row=current_row, column=1, value=f"  {step}")
                cell.alignment = LEFT
                cell.border = THIN_BORDER
                ws.row_dimensions[current_row].height = 18
                current_row += 1

            # HACCP CCP
            if sop.ccp_checkpoints:
                ws.cell(row=current_row, column=1, value="[HACCP CCP]").font = Font(bold=True, color="C00000")
                ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor="FFE7E7")
                ws.cell(row=current_row, column=1).border = CCP_BORDER
                current_row += 1
                for ccp in sop.ccp_checkpoints:
                    ws.merge_cells(f"A{current_row}:E{current_row}")
                    cell = ws.cell(row=current_row, column=1, value=f"  ▶ {ccp}")
                    cell.fill = PatternFill("solid", fgColor="FFE7E7")
                    cell.border = CCP_BORDER
                    cell.font = Font(color="C00000")
                    cell.alignment = LEFT
                    current_row += 1

            # Storage + allergy
            ws.cell(row=current_row, column=1, value="[보관]").fill = ACCENT_FILL
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.cell(row=current_row, column=1).border = THIN_BORDER
            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws.cell(row=current_row, column=2, value=sop.storage_instruction).border = THIN_BORDER
            ws.cell(row=current_row, column=2).alignment = LEFT
            current_row += 1

            if sop.allergen_note:
                ws.cell(row=current_row, column=1, value="[알레르기]").fill = ORANGE_FILL
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                ws.cell(row=current_row, column=1).border = THIN_BORDER
                ws.merge_cells(f"B{current_row}:E{current_row}")
                ws.cell(row=current_row, column=2, value=sop.allergen_note).border = THIN_BORDER
                ws.cell(row=current_row, column=2).alignment = LEFT
                current_row += 1

            current_row += 1  # blank spacer

        for i, w in enumerate([14, 20, 18, 18, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Helpers ─────────────────────────────────────────────────────────────

    def _load_menu_index(self) -> dict:
        try:
            import pandas as pd
            from pathlib import Path
            data_dir = Path(__file__).parent.parent / "data"
            menu_df = pd.read_csv(data_dir / "menu_db.csv").set_index("menu_id")
            return menu_df["menu_name_ko"].to_dict()
        except Exception:
            return {}
